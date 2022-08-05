# Importing the openpyxl package
import openpyxl

path_to_acetylome_data = "Acetylome_Results_Contaminants_Removed.xlsx"
path_to_total_proteomics_data = "Total_Proteomics_Contaminants_FALSE_Table_Sorted.xlsx"
path_to_acetylome_data_write_out = "Acetylome_Results_Contaminants_Removed_Write_Out.xlsx"


# Conducts a binary search through the total proteomics data set searching for an accession that matches the acetylome
# accession passed as a parameter. Returns an array of the accession and its corresponding p53 OFF WCL, p53 OFF nuclear,
# p53 OFF cytoplasm, p53 ON WCL, p53 ON nuclear, p53 ON cytoplasm averages if a matching accession is found. Returns -1
# if no match is found.
def binary_search_total_proteome(proteome_array, low, high, accession):
    if high >= low:
        mid = (high + low) // 2

        if proteome_array[mid][0] == str(accession):
            return_list = [None for x in range(7)]
            for i in range(7):
                return_list[i] = proteome_array[mid][i]
            return return_list

        elif proteome_array[mid][0] > str(accession):
            return binary_search_total_proteome(proteome_array, low, mid - 1, accession)

        else:
            return binary_search_total_proteome(proteome_array, mid + 1, high, accession)

    else:
        return -1

# Passes each accession for a given acetylome identifier to the binary search method until a match is found. Returns
# the result of the binary search.
def find_matching_accession(accession):
    matched_list = None
    matching_accession_found = False
    for i in range(len(accession)):
        if matching_accession_found is False:
            matched_list = binary_search_total_proteome(total_proteome_array, 0, len(total_proteome_array) - 1,
                                                        accession[i])
            if matched_list != -1:
                matching_accession_found is True
                break
    return matched_list


# Creates and returns a 2D array of the acetylome accessions for each acetylome identifier in the
# Acetylome_Results_Contaminants_Removed.xlsx Excel workbook.
def create_acetylome_accession_array():
    # Loads the Acetylome_Results_Contaminants_Removed Excel workbook
    workbook = openpyxl.load_workbook(path_to_acetylome_data, data_only=True)

    # Sets the active sheet as the first sheet in the Acetylome_Results_Contaminants_Removed.xlsx Excel workbook
    active_sheet = workbook.active

    max_row = active_sheet.max_row

    i = max_row - 1
    j = 6

    accession_array = [[None for x in range(j)] for y in range(i)]

    for row in active_sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=6):
        for cell in row:
            accession_array[cell.row - 2][cell.column - 1] = cell.value

    return accession_array

# Creates and returns a 2D array of the total proteome accessions and their corresponding p53 OFF WCL, p53 OFF nuclear,
# p53 OFF cytoplasm, p53 ON WCL, p53 ON nuclear, p53 ON cytoplasm averages.
def create_total_proteome_array():
    # Loads the Total_Proteomics_Contaminants_FALSE_Table_Sorted Excel workbook
    workbook = openpyxl.load_workbook(path_to_total_proteomics_data, data_only=True)

    # Sets the active sheet as the first sheet in the Total_Proteomics_Contaminants_FALSE_Table_Sorted Excel workbook
    active_sheet = workbook.active

    max_row = active_sheet.max_row

    i = max_row - 1
    j = 7

    total_proteome_array = [[None for x in range(j)] for y in range(i)]

    for row in active_sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=38):
        for cell in row:
            if cell.column == 1:
                total_proteome_array[cell.row - 2][cell.column - 1] = cell.value
            elif 33 <= cell.column <= 38:
                total_proteome_array[cell.row - 2][cell.column - 32] = cell.value
    return total_proteome_array


acetylome_accession_array = create_acetylome_accession_array()
total_proteome_array = create_total_proteome_array()

# Loops through the acetylome accession array passing each index to the find_matching_accession() method. Creates
# a new array of the returned values and then loops through the new array writing out each index to the
# Acetylome_Results_Contaminants_Removed_Write_Out.xlsx Excel workbook. If no matching accession was found in the total
# proteome data set, values of #N/A are entered into the p53 OFF WCL, p53 OFF nuclear, p53 OFF cytoplasm, p53 ON WCL,
# p53 ON nuclear, p53 ON cytoplasm averages columns.
def match_accessions():
    final_matched_array = [None for x in range(len(acetylome_accession_array))]
    for i in range(len(acetylome_accession_array)):
        final_matched_array[i] = find_matching_accession(acetylome_accession_array[i])
    # Loads the Acetylome_Results_Contaminants_Removed Excel workbook
    workbook = openpyxl.load_workbook(path_to_acetylome_data_write_out, data_only=True)

    # Sets the active sheet as the first sheet in the Acetylome_Results_Contaminants_Removed.xlsx Excel workbook
    active_sheet = workbook.active

    for row in active_sheet.iter_rows(min_row=2, max_row=5700, min_col=39, max_col=44):
        for cell in row:
            if final_matched_array[cell.row - 2] != -1:
                cell.value = final_matched_array[cell.row - 2][cell.column - 38]
            else:
                cell.value = "#N/A"
    workbook.save(path_to_acetylome_data_write_out)

match_accessions()
